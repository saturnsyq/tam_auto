#!/usr/bin/python3
#coding=utf-8

import imaplib
import email
from email.parser import Parser
from email.header import decode_header
from email.utils import parseaddr
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.utils import COMMASPACE,formatdate
from email import encoders

from bs4 import BeautifulSoup
import pymysql
import importlib
import sys
import logging
import getpass
import time
import xlrd
import os
import xlwt
from openpyxl import load_workbook
import nltk
import unicodedata
import re
import datetime

import subprocess

import taurus
import vendorMail
import metadataProcessor

from imp import reload
import requests
from concurrent.futures import ThreadPoolExecutor
import mailProcessor_config
import tamCommonLib
from openpyxl import Workbook

## comment variables ##########################################
## time zone
mysql_host = 'ud094661c879c59fa6e9e'
mysql_user = '***'
mysql_password = '***'
mysql_default_db = '***'
work_path = './'
## mysql configuration
importlib.reload(sys)

##logging configuration
logging.basicConfig(level=logging.DEBUG,
                format='%(asctime)s %(filename)s[line:%(lineno)3d] %(levelname)s %(message)s',
                datefmt='%Y-%m-%d %a %H:%M:%S',
                filename=work_path+'/mailProcessor.log',
                filemode='w')

###############################################################
months = {'jan':'01','feb':'02','mar':'03','apr':'04','may':'05','jun':'06','jul':'07','aug':'08','sep':'09','oct':'10','nov':'11','dec':'12'}
allow_addresses = ['@amazon.com','@a9.com']

helpContent = '''
<html>
<head></head>
<body>
<p>Hello %s,</p>
<p>Thank you for emailing TAM AUTO system.</p>
<p>TAM AUTO system is a powerful tool owned by CN TAM Team. We developed this platform to fulfill all kinds of data requests submitted via amazon internal mails.</p>
<p>Our vision is <font style="font-style:italic;">'Anywhere, any data, in 60 seconds'.</font> </p>
<p>
Imagine you are visiting a vendor, and you need data including daily paid units and conversion rate for a Kindle ASIN urgently. You can just send a mail to cn-tam-auto@ by your mobile and all data requested will appear on your screen in 60 seconds!   
</p>
<p/>
<p>Interested? Try below links now.</p>
<p>
    <table style="white-space:nowrap;"> 
        %s
    </table>
</p>
<p/>
<p>Stay tuned, more to come!</p>
<p/>
<p>Please refer the <a href="https://w.amazon.com/bin/view/CN_TAM_PATTERN/">wiki</a> to get help for some features.</p>
<p>Any questions or suggestions, please cut a SIM:</p>
<p><a href="http://tiny.amazon.com/rmk7yb14">http://tiny.amazon.com/rmk7yb14</a></p>
<p>Or contact chenmiao@ / nanmeng@ / yongqis@ directly.</p>
<p/>
<p>Regards,</p>
<p>CN TAM Team</p>
</body>
</html>
'''

bodyMessage = '''
<html>
<head></head>
<body>
<p>Hello %s,</p>
<p>Please see the data in the attachment.</p>
<p/>
<p/>
<p><font style="font-weight:bold;">Notice: DO NOT Forward this mail.</font></p>
<p><font style="font-weight:bold;">It's only for internal use.</font></p>
</body>
</html>
'''

# 邮件的Subject或者Email中包含的名字都是经过编码后的str，要正常显示，就必须decode
def decode_str(s):
    value, charset = decode_header(s)[0]
    # decode_header()返回一个list，因为像Cc、Bcc这样的字段可能包含多个邮件地址，所以解析出来的会有多个元素。上面的代码我们偷了个懒，只取了第一个元素。
    if charset:
        value = value.decode(charset)
    return value


# 文本邮件的内容也是str，还需要检测编码，否则，非UTF-8编码的邮件都无法正常显示
def guess_charset(msg):
    charset = msg.get_charset()
    if charset is None:
        content_type = msg.get('Content-Type', '').lower()
        pos = content_type.find('charset=')
        if pos >= 0:
            charset = content_type[pos + 8:].strip()
    return charset


def send_mail(fro, to, subject, text, files=[], mtype='plain'):
    #assert type(server) == dict
    #assert type(to) == list
    #assert type(files) == list

    msg = MIMEMultipart()
    msg['From'] = fro
    msg['Subject'] = subject
    msg['To'] = ','.join(to)  # COMMASPACE==', '
    #msg['Date'] = formatdate(localtime=True)
    msg.attach(MIMEText(text,mtype))

    for file in files:
        part = MIMEBase('application', 'octet-stream')  # 'octet-stream': binary data
        part.set_payload(open(file, 'rb').read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', 'attachment; filename="%s"' % os.path.basename(file))
        msg.attach(part)

    import smtplib
    smtp = smtplib.SMTP('mail-relay.amazon.com')
    #smtp.login(server['user'], server['passwd'])
    smtp.sendmail(fro, to, msg.as_string())
    smtp.close()

#if sql end with ';', remove it
def sql_correct(sql):
    if sql is None or len(sql)==0: return sql
    new_sql=sql.strip()
    if new_sql[-1] in ';；':
        new_sql=new_sql[:-1]
    return new_sql

def checkAsinOnsite(asin):
    r = requests.get(url='https://www.amazon.cn/s', params={'field-keywords': asin})
    exists = True
    if r.text.find(u'没有找到任何与')>0 and r.text.find(u'相关的商品') and r.text.find(u'请尝试以下内容'):
        exists = False
    return exists

# 解析邮件与构造邮件的步骤正好相反
def process(msg, msgid, name):
    subject = ""
    for header in ["From", "To", "Subject", "Cc"]:
        value = msg.get(header, "")
        if value:
            if header == "Subject":
                value = decode_str(value)
                # else:
                #     hdr, addr =parseaddr(value)
                #     name = decode_str(hdr)
                #     value = u"%s <%s>" % (name, addr)
        print("%s:%s" % (header, value))
        if header == "Subject": subject = value
    mail_replied = False
    for part in msg.walk():
        fileName = part.get_filename()
        contentType = part.get_content_type()
        mycode = part.get_content_charset();
        # 保存附件
        if contentType.find('image/') != -1: continue
        if fileName:
            data = part.get_payload(decode=True)
            fname = decode_str(fileName)
            fEx = open("%s" % (fname), 'wb')
            fEx.write(data)
            fEx.close()
            if fname.find('.xls') >=0:
                #read the excel file
                data = xlrd.open_workbook(fname)
                table = data.sheet_by_index(0)
                nrows = table.nrows
                ncols = table.ncols
                keycol_pos = {}
                asin_col_name =""
                schedule_col_name =""
                title_col_name =""
                vm_col_name =""
                k_code_name =""
                for j in range(ncols):
                    temp = str(table.cell(0,j).value).strip().lower()
                    keycol_pos[temp] = j
                    if temp.find('asin') >= 0:
                        if asin_col_name!='asin': asin_col_name = temp
                    elif temp.find('schedule')>= 0 or temp.find('time')>= 0:
                        if schedule_col_name!='time': schedule_col_name = temp
                    elif temp.find('title')>= 0:
                        if title_col_name!='title': title_col_name = temp
                    elif temp.find('vm')>= 0:
                        if vm_col_name!='vm': vm_col_name = temp
                    elif temp.find('code')>= 0:
                        if k_code_name!='k_code': k_code_name = temp
                #get and process the data
                # if keycol_pos.get('asin') is None or keycol_pos.get('schedule') is None or keycol_pos.get('title') is None:
                if len(asin_col_name)==0 or len(schedule_col_name)==0 or len(title_col_name)==0:
                    logging.warning('asin or schedule or title columns don\'t exist in the attached excel file.\nSo skip it')
                    send_mail('cn-tam-auto@amazon.com', [msg.get('From', "")], 'ERROR:' + subject, 'asin or schedule or title columns don\'t exist in the attached excel file.So skip it', [])
                    return
                conn = pymysql.connect(host=mysql_host, user=mysql_user, passwd=mysql_password,
                                       db=mysql_default_db, port=3306, local_infile=1, charset='utf8',
                                       cursorclass=pymysql.cursors.DictCursor)
                cursor = conn.cursor()
                #newWB = copy(xlrd.open_workbook('promotion_data_sample.xlsx', formatting_info=True))
                #newWB = copy(xlrd.open_workbook('promotion_data_sample.xlsx'))
                newWB = load_workbook('promotion_data_sample.xlsx')
                newWs = newWB.get_active_sheet()
                for i in range(1,nrows):
                    asin = str(table.cell(i,keycol_pos[asin_col_name]).value).strip()
                    schedule = str(table.cell(i, keycol_pos[schedule_col_name]).value)
                    title = str(table.cell(i, keycol_pos[title_col_name]).value)
                    pub_code = ''
                    if len(k_code_name)>0:
                        pub_code = str(table.cell(i, keycol_pos[k_code_name]).value)
                    _=newWs.cell(row=i+1, column=1, value=schedule)
                    _=newWs.cell(row=i+1, column=2, value=asin)
                    _=newWs.cell(row=i+1, column=3, value=title)
                    _=newWs.cell(row=i+1, column=4, value=pub_code)
                    if len(vm_col_name)>0:
                        _=newWs.cell(row=i+1, column=7, value=str(table.cell(i, keycol_pos[vm_col_name]).value))
                    cursor.execute("select case when v.onsite='y' or v.onsite='Y' then 'y' else ab.is_os end as onsite_n,v.iacipc,v.vendor_code,v.vm,v.tam,v.m_q_pub,v.tam_ops from tam_db.visibility v,tam_db.asin_base ab where v.asin=ab.asin and v.asin='%s'" % asin)
                    row = cursor.fetchone()
                    if asin is None or len(asin)==0 or row is None:
                        cursor.execute("select vm,max(tam) as tam,max(m_q_pub) as m_q_pub,max(tam_ops) as tam_ops,count(1) as cnt from tam_db.visibility where vendor_code='%s' and vm is not null and length(vm)>0 group by vm order by count(1) desc" % pub_code)
                        row = cursor.fetchone()
                        if row is None: continue
                    if row.get('vendor_code') is not None:
                        pub_code = row['vendor_code']
                    _=newWs.cell(row=i+1, column=4, value=pub_code)
                    if row.get('onsite_n') is not None:
                        temp_n = row['onsite_n']
                        # if temp_n=='n' and checkAsinOnsite(asin):
                        #    temp_n = 'y'
                        _=newWs.cell(row=i+1, column=5, value=temp_n)
                    if row.get('iacipc') is not None:
                        _=newWs.cell(row=i+1, column=6, value=row['iacipc'])
                    if row.get('vm') is not None:
                        _=newWs.cell(row=i+1, column=7, value=row['vm'])
                    if row.get('tam_ops') is not None:
                        _=newWs.cell(row=i+1, column=8, value=row['tam_ops'])
                    if row.get('tam') is not None:
                        _=newWs.cell(row=i+1, column=9, value=row['tam'])
                    if row.get('m_q_pub') is not None:
                        _=newWs.cell(row=i+1, column=10, value=row['m_q_pub'])
                newWB.save('./promotion_data.xlsx')
                send_mail('cn-tam-auto@amazon.com',[msg.get('From', "")],'RE:'+subject,'Please see the data in the attachment.',['./promotion_data.xlsx'])
                fro = re.findall(r'([a-zA-Z0-9_\-\.]+@[a-zA-Z0-9]+\.[a-zA-Z0-9]+)', msg.get('From', ""))[0]
                rec_time = re.findall(r'\w+, (\d{1,2}) (\w+) (\d{4}) (\d{2}:\d{2}:\d{2})', msg.get("Date", ""))[0]
                rec_str =  str(datetime.datetime.strptime("%s-%s-%s %s" % (rec_time[2], months[rec_time[1].lower()], rec_time[0], rec_time[3]),"%Y-%m-%d %H:%M:%S") + datetime.timedelta(hours=8))
                cursor.execute("insert into accesslog.mail_audit(msgid,fro,receive,subject,type) values (%s,'%s','%s','%s','%s')" % ( msgid, fro, rec_str, subject, '#DOTD'))
                conn.commit()
                cursor.close()
                conn.close()
                #row_data = [str(table.cell(i,j).value) for j in range(ncols)]
                #print('|'.join(row_data))
                if os.path.exists(fname): os.remove(fname)
                if os.path.exists('./promotion_data.xlsx'): os.remove('./promotion_data.xlsx')
                mail_replied = True

        elif contentType == 'text/plain'or contentType == 'text/html':
            # 保存正文
            content = part.get_payload(decode=True)
            charset = guess_charset(part) #or msg
            if charset:
                content = content.decode(charset)
            #print('%sText: %s' % ('  ' * indent, content + '...'))
            soup = BeautifulSoup(content,'lxml')
            table = soup.find('table')
            if table is not None:
                line_no = 0
                asin_index = -1
                data = []
                for tr in table.findAll('tr'):
                    line_no += 1
                    col_no = 0
                    asin = ""
                    for td in tr.findAll('td'):
                        col_no += 1
                        if line_no == 1:
                            if td.getText().strip().lower() == "asin" or td.getText().strip().lower() == "kasin":
                                asin_index = col_no
                        else:
                            if asin_index == col_no:
                                asin = td.getText().strip()
                    comments = td.getText().strip() # last column for comments
                    if line_no!=1 and len(comments)>0 and len(asin)>0:
                        data.append((asin,comments,msgid))
                if len(data) >0:
                    print("\nstart to load the data into mysql database.")
                    conn = pymysql.connect(host=mysql_host, user=mysql_user, passwd=mysql_password,
                                           db=mysql_default_db, port=3306, local_infile=1, charset='utf8',
                                           cursorclass=pymysql.cursors.DictCursor)
                    cursor = conn.cursor()
                    for line in data:
                        print(line[0]+":"+line[1])
                        sql = "delete from tam_db.tam_auto_data where asin='%s' and msgid=%s" % (line[0],line[2])
                        cursor.execute(sql)
                        sql = "insert into tam_db.tam_auto_data(asin,comments,msgid) values ('%s','%s',%s)" % line
                        cursor.execute(sql)
                    conn.commit()
                    print("Finish to load the data.")
    if mail_replied == False:
        send_mail('cn-tam-auto@amazon.com', [msg.get('From', "")], 'ERROR:' + subject,'Only fixed layout excel attachement can be processed by this pattern.\nSo skip it', [])


def newSheetWithSQL(sql,sheet,cursor):
    dateFormat = xlwt.XFStyle()
    dateFormat.num_format_str = 'yyyy/mm/dd'
    cursor.execute(sql)
    results = cursor.fetchall()
    fields = cursor.description
    # 写上字段信息
    for field in range(0, len(fields)):
        sheet.write(0, field, fields[field][0])
    # 获取并写入数据段信息
    row = 1
    col = 0
    for row in range(1, len(results) + 1):
        for col in range(0, len(fields)):
            if results[row - 1].get(fields[col][0]) is not None:
                if isinstance(results[row - 1].get(fields[col][0]), datetime.date):
                    #sheet.write(row, col, results[row - 1].get(fields[col][0]), dateFormat)
                    sheet.write(row, col, str(results[row - 1].get(fields[col][0])))
                else:
                    sheet.write(row, col, results[row - 1].get(fields[col][0]))

def processPubScan(msg,msgid,name):
    subject = ""
    for header in ["From", "To", "Subject", "Cc"]:
        value = msg.get(header, "")
        if value:
            if header == "Subject":
                value = decode_str(value)
        #print("%s:%s" % (header, value))
        if header == "Subject": subject = value
    for part in msg.walk():
        # fileName = part.get_filename()
        contentType = part.get_content_type()
        # mycode = part.get_content_charset();
        # 保存附件
        if contentType.find('image/') != -1: continue
        if contentType == 'text/plain' or contentType == 'text/html':
            # 保存正文
            content = part.get_payload(decode=True)
            charset = guess_charset(part)  # or msg
            if charset:
                content = content.decode(charset)
            clearText = BeautifulSoup(content, 'lxml').get_text().strip().replace(u'\xa0', u' ')
            pubs = re.findall(r'(\w{3}CN)',clearText.upper())
            if len(pubs) == 0:
                send_mail('cn-tam-auto@amazon.com', [msg.get('From', "")], 'RE:' + subject,
                              'No pub was recognized\n\nPlease make sure you have provided the valid pub list.\n\nFor example:\nYKJCN\nSDKCN', [])
                continue
            in_condition = "','".join(pubs)
            conn = pymysql.connect(host=mysql_host, user=mysql_user, passwd=mysql_password,
                                   db=mysql_default_db, port=3306, local_infile=1, charset='utf8',
                                   cursorclass=pymysql.cursors.DictCursor)
            cursor = conn.cursor()
            wk_data = datetime.datetime.strftime(datetime.datetime.now(),'_%Ywk%W')
            mtype,key_data = '',''
            try:
                #cursor.execute("select distinct pubcode from tam_db.vendor_LTD where pubcode in ('%s')" % in_condition)
                cursor.execute("select distinct pubcode from tam_db.vendor_LTD where pubcode in ('%s')" % pubs[0])
                pubs_list = []
                pubs = []
                for pub in cursor.fetchall():
                    workbook = xlwt.Workbook()
                    #retrieve FL deal data
                    sheet = workbook.add_sheet("FL Deal", cell_overwrite_ok=True)
                    sql = "select * from tam_db.vendor_fl_deal where pubcode='%s' order by onsite_date desc" % pub['pubcode']
                    newSheetWithSQL(sql,sheet,cursor)
                    # Low_CR data
                    sheet = workbook.add_sheet("Low_CR", cell_overwrite_ok=True)
                    sql = "select * from tam_db.vendor_LTD where pubcode='%s' order by isnull(cr_t30),cr_t30,onsite_date desc" % pub['pubcode']
                    newSheetWithSQL(sql, sheet, cursor)
                    # No_GV data
                    sheet = workbook.add_sheet("No_GV", cell_overwrite_ok=True)
                    sql = "select * from tam_db.vendor_LTD where pubcode='%s' order by isnull(gv_t30),gv_t30,onsite_date desc" % pub['pubcode']
                    newSheetWithSQL(sql, sheet, cursor)
                    # Supp_Title data
                    sheet = workbook.add_sheet("Supp_Title", cell_overwrite_ok=True)
                    sql = "select * from tam_db.vendor_LTD where kindle_suppression_state='COMPLETELY_SUPPRESSED' and pubcode='%s' order by onsite_date desc" % pub['pubcode']
                    newSheetWithSQL(sql, sheet, cursor)
                    # No_Deal data
                    sheet = workbook.add_sheet("No_Deal", cell_overwrite_ok=True)
                    sql = "select * from tam_db.vendor_LTD where deal_flag='n' and pubcode='%s' order by onsite_date desc" %  pub['pubcode']
                    newSheetWithSQL(sql, sheet, cursor)
                    # Inactive_Title data
                    sheet = workbook.add_sheet("Inactive_Title", cell_overwrite_ok=True)
                    sql = "select * from tam_db.vendor_LTD where pubcode='%s' order by isnull(pu_t30),pu_t30,onsite_date desc" % pub['pubcode']
                    newSheetWithSQL(sql, sheet, cursor)
                    # Deal data
                    sheet = workbook.add_sheet("Deal", cell_overwrite_ok=True)
                    sql = "select asin,title_name,onsite_day,new_price,promotion_title, \
                               start_day,end_day,primary_promotion_name,secondary_promotion_name, \
                              prior_units,campaign_units,post_units,prior_gms,campaign_gms,post_gms \
                         from tam_db.vendor_deal_uplift where digital_pubcode='%s' order by onsite_day desc,asin,start_day desc" % pub['pubcode']
                    newSheetWithSQL(sql, sheet, cursor)
                    #pubs_list.append(pub['pubcode'] + '.xls')
                    pubs.append(pub['pubcode'])
                    workbook.save(pub['pubcode'] + wk_data + '.xls')
                    # gzip the file if it is greater than 5M
                    if os.stat(pub['pubcode'] +  wk_data + '.xls').st_size / (1024 * 1024) > 5:
                        os.system("gzip " + pub['pubcode'] + wk_data + '.xls')
                        if os.path.exists(pub['pubcode'] + wk_data + '.xls.gz'):
                            pubs_list.append(pub['pubcode'] + wk_data + '.xls.gz')
                    else:
                        pubs_list.append(pub['pubcode'] + wk_data + '.xls')
                #
                key_data = ','.join(pubs)[:250]
                send_mail('cn-tam-auto@amazon.com', [msg.get('From', "")], 'RE:' + subject, "Please see the attached data.",pubs_list)
                for ff in pubs_list:
                    if os.path.exists(ff):
                        os.remove(ff)
            except Exception as e:
                errMsg = repr(e)
                send_mail('cn-tam-auto@amazon.com', [msg.get('From', "")], 'Error:' + subject, errMsg, [])
            finally:
                cursor.close()
                conn.close()

            return {'key_data': key_data, 'mtype': '#PubScan'}

def processTPOS(msg,msgid,name):
    subject = ""
    for header in ["From", "To", "Subject", "Cc"]:
        value = msg.get(header, "")
        if value:
            if header == "Subject":
                value = decode_str(value)
        #print("%s:%s" % (header, value))
        if header == "Subject": subject = value
    for part in msg.walk():
        # fileName = part.get_filename()
        contentType = part.get_content_type()
        # mycode = part.get_content_charset();
        # 保存附件
        if contentType.find('image/') != -1: continue
        if contentType == 'text/plain' or contentType == 'text/html':
            # 保存正文
            conn = pymysql.connect(host=mysql_host, user=mysql_user, passwd=mysql_password,
                                   db=mysql_default_db, port=3306, local_infile=1, charset='utf8',
                                   cursorclass=pymysql.cursors.DictCursor)
            cursor = conn.cursor()
            try:
                workbook = xlwt.Workbook()
                #retrieve Deal Strong data
                sheet = workbook.add_sheet("Deal_Strong", cell_overwrite_ok=True)
                sql = '''
select 
temp1.snapshot_day
,temp1.asin KASIN
,ab.pasin PASIN
,temp1.title_name
,temp1.pubcode
,ab.vm_name
,ab.team_name
,ab.category_name
,ab.sub_category_name
,temp1.onsite_day as K_onsite_day
,ab.p_release_date as P_release_day
-- ,temp1.PROMOTION_TITLE
,ad.primary_promotion_name
,ad.secondary_promotion_name
,temp1.d_s_date as Deal_Start
,temp1.d_e_date as Deal_End
,datediff(temp1.d_e_date,temp1.d_s_date) as Deal_Duration
-- ,ad.gms as Deal_GMS
,temp1.d_d_gms as Deal_Daily_GMS
,round(ab.k_t30_gms/30,1) as K_T30_Daily_GMS
,temp1.d_gms as  K_LTD_Daily_GMS
,temp1.d_times as Deal_Uplift_Ratio
,ad.deal_price as Deal_OP
,ab.our_price as K_OP
,ab.list_price as K_LP
,temp1.is_new
from
(
select a1.*,'N' as is_new 
from tam_db.deal_strong a1 
where a1.snapshot_day = ( select max(snapshot_day) from tam_db.deal_strong ) 
and exists 
( select 1 from tam_db.deal_strong a2 where a1.asin=a2.asin and a2.snapshot_day= 
( select max(snapshot_day) from tam_db.deal_strong 
where snapshot_day!=(select max(snapshot_day) from tam_db.deal_strong)) ) 
union 
select a1.*,'Y' as is_new 
from tam_db.deal_strong a1 
where a1.snapshot_day = ( select max(snapshot_day) from tam_db.deal_strong ) 
and not exists 
( select 1 from tam_db.deal_strong a2 where a1.asin=a2.asin and a2.snapshot_day= 
( select max(snapshot_day) from tam_db.deal_strong 
where snapshot_day!=(select max(snapshot_day) from tam_db.deal_strong)) )
) temp1
left join tam_db.asin_base ab
on temp1.asin = ab.asin
left join tam_db.asin_deal ad
on temp1.asin = ad.asin and temp1.d_s_date = ad.start_date
where ab.kindle_suppression_state <> 'COMPLETELY_SUPPRESSED'
order by is_new desc, Deal_Start desc
                '''
                newSheetWithSQL(sql,sheet,cursor)
                #retrieve Dual High data
                sheet = workbook.add_sheet("Dual_high", cell_overwrite_ok=True)
                sql = '''
select
temp1.snapshot_day
,temp1.kasin as KASIN
,ab.pasin as PASIN
,temp1.name as Title_Name
,temp1.pubcode
,ab.vm_name
,ab.team_name
,ab.category_name
,ab.sub_category_name
,temp1.k_onsite_date as K_onsite_day
,ab.p_release_date as P_release_day
,ab.our_price as K_OP
,ab.list_price as K_LP
,round(ab.k_t30_gvc*10000,3) as K_T30_GVC
,case when temp1.k_op = 0 then 'N'
 else 'Y'
 end K_Rule
,ab.p_our_price as P_OP
,ab.p_list_price as P_LP
,round(ab.p_t30_gvc*10000,3) as P_T30_GVC
,case when temp1.p_op = 0 then 'N'
 else 'Y'
 end P_Rule
,ad.primary_promotion_name
,ad.secondary_promotion_name
,temp1.start_date as Deal_Start
,temp1.end_date as Deal_End
,datediff(temp1.end_date,temp1.start_date) as Deal_Duration
,ad.gms as Deal_GMS
,round(ad.gms/datediff(temp1.end_date,temp1.start_date),1) as Deal_Daily_GMS
,round(ab.k_t30_gms/30,1) as K_T30_Daily_GMS
,round(ab.k_ltd_gms/ab.k_onsite_duration,1) as LTD_Daily_GMS
,temp1.is_new
from
(
select a1.*,'N' as is_new 
from tam_db.dual_high a1 
where a1.snapshot_day = ( select max(snapshot_day) from tam_db.dual_high ) 
and exists 
( select 1 from tam_db.dual_high a2 where a1.kasin=a2.kasin and a2.snapshot_day= 
( select max(snapshot_day) from tam_db.dual_high 
where snapshot_day!=(select max(snapshot_day) from tam_db.dual_high)) ) 
union 
select a1.*,'Y' as is_new 
from tam_db.dual_high a1 
where a1.snapshot_day = ( select max(snapshot_day) from tam_db.dual_high ) 
and not exists 
( select 1 from tam_db.dual_high a2 where a1.kasin=a2.kasin and a2.snapshot_day= 
( select max(snapshot_day) from tam_db.dual_high 
where snapshot_day!=(select max(snapshot_day) from tam_db.dual_high)) )
) temp1
left join tam_db.asin_base ab
on temp1.kasin = ab.asin
left join tam_db.asin_deal ad
on temp1.kasin = ad.asin and temp1.start_date = ad.start_date
where ab.kindle_suppression_state <> 'COMPLETELY_SUPPRESSED'
order by is_new desc, K_T30_GVC desc
                '''
                newSheetWithSQL(sql, sheet, cursor)
                #retrieve UpStream data
                sheet = workbook.add_sheet("Upstream", cell_overwrite_ok=True)
                sql = '''
select 
temp1.snapshot_day
,temp1.kasin as KASIN
,ab.pasin as PASIN
,temp1.name as Title_Name
,temp1.pubcode
,ab.vm_name
,ab.team_name
,ab.category_name
,ab.sub_category_name
,temp1.k_onsite_date as K_onsite_day
,ab.p_release_date as P_release_day
,temp1.k_t7_gvc_after
,temp1.k_gv_before
,temp1.k_gv_after
,temp1.k_rate as K_GV_Uplift
,case when temp1.k_t7_gvc_after = 0 then 'N'
 else 'Y'
 end K_Rule
,temp1.p_t7_gvc_after
,temp1.p_gv_before
,temp1.p_gv_after
,temp1.p_rate as P_GV_Uplift
,case when temp1.p_t7_gvc_after = 0 then 'N'
 else 'Y'
 end P_Rule
,ad.primary_promotion_name
,ad.secondary_promotion_name
,temp1.start_date as Deal_Start
,temp1.end_date as Deal_End
,datediff(temp1.end_date,temp1.start_date) as Deal_Duration
,ad.gms as Deal_GMS
,round(ad.gms/datediff(temp1.end_date,temp1.start_date),1) as Deal_Daily_GMS
,round(ab.k_t30_gms/30,1) as K_T30_Daily_GMS
,round(ab.k_ltd_gms/ab.k_onsite_duration,1) as LTD_Daily_GMS
,ab.our_price as K_OP
,ab.list_price as K_LP
,ab.p_our_price as P_OP
,ab.p_list_price as P_LP
,temp1.is_new
from
(
select a1.*,'N' as is_new 
from tam_db.upstream a1 
where a1.snapshot_day = ( select max(snapshot_day) from tam_db.upstream ) 
and exists 
( select 1 from tam_db.upstream a2 where a1.kasin=a2.kasin and a2.snapshot_day= 
( select max(snapshot_day) from tam_db.upstream 
where snapshot_day!=(select max(snapshot_day) from tam_db.upstream)) ) 
union 
select a1.*,'Y' as is_new 
from tam_db.upstream a1 
where a1.snapshot_day = ( select max(snapshot_day) from tam_db.upstream ) 
and not exists 
( select 1 from tam_db.upstream a2 where a1.kasin=a2.kasin and a2.snapshot_day= 
( select max(snapshot_day) from tam_db.upstream 
where snapshot_day!=(select max(snapshot_day) from tam_db.upstream)) )
) temp1
left join tam_db.asin_base ab
on temp1.kasin = ab.asin
left join tam_db.asin_deal ad
on temp1.kasin = ad.asin and temp1.start_date = ad.start_date
where ab.kindle_suppression_state <> 'COMPLETELY_SUPPRESSED'
order by temp1.is_new desc, temp1.start_date desc
                '''
                newSheetWithSQL(sql, sheet, cursor)
                #deadlock pattern
                sql='''
                SELECT dr.asin as kasin,
                   ab.title_name as title_name,
        ab.band as actual_band,
        dr.price as recent_price,
        dr.pu as recent_pu,
        dr.gv as recent_gv, 
        dr.days as days,
        concat(round(dr.pu*100/dr.gv,2),'%') as cr, 
        is_ku_flag ,
        ab.vm_name, 
        list_price as DLP, 
        p_our_price,
        concat(round(dr.price*100/p_our_price,2),'%') as OP_ratio,
        tam_db.asin_base.onsite_date as os_date 
        from tam_db.price_duration_recent dr
        
inner join hera.all_bands ab on ab.asin  = dr.asin
inner join (SELECT asin, min(price) as price FROM tam_db.price_duration group by asin) aa on aa.asin = dr.asin and dr.price = aa.price
inner join tam_db.asin_base on tam_db.asin_base.asin = dr.asin

where days>=60 and dr.price <10
and dr.asin in (select asin from hera.all_bands where DIGITAL_PUBCODE <> 'YWGCN' )
and dr.asin in (SELECT asin FROM tam_db.price_match where entry_day_local = (select max(entry_day_local) from tam_db.price_match))
and dr.asin in (SELECT asin FROM tam_db.price_duration group by asin having count(distinct price)>2)
and dr.price/tam_db.asin_base.p_our_price < 0.2
order by dr.price
                '''
                sheet = workbook.add_sheet("Deadlock", cell_overwrite_ok=True)
                newSheetWithSQL(sql, sheet, cursor)

                ff_name = "TAM_Pattern_OS_Part.xls"
                workbook.save(ff_name)
                body_text = '''
              <p>See the attached data.</p>
              <p>Please refer the <a href="https://w.amazon.com/bin/view/CN_TAM_PATTERN/">wiki</a> to get help for some features.</p>
              '''
                send_mail('cn-tam-auto@amazon.com', [msg.get('From', "")], 'TAM Pattern OS Part', body_text,[ff_name],'html')
                if os.path.exists(ff_name):
                    os.remove(ff_name)
                mtype = "#TPOS"
                key_data = ''
                fro = re.findall(r'([a-zA-Z0-9_\-\.]+@[a-zA-Z0-9]+\.[a-zA-Z0-9]+)', msg.get('From', ""))[0]
                rec_time = re.findall(r'\w+, (\d{1,2}) (\w+) (\d{4}) (\d{2}:\d{2}:\d{2})', msg.get("Date", ""))[0]
                rec_str = str(datetime.datetime.strptime(
                    "%s-%s-%s %s" % (rec_time[2], months[rec_time[1].lower()], rec_time[0], rec_time[3]),
                    "%Y-%m-%d %H:%M:%S") + datetime.timedelta(hours=8))
                cursor.execute("insert into accesslog.mail_audit(msgid,fro,receive,subject,type,key_data) values (%s,'%s','%s','%s','%s','%s')" % (
                        msgid, fro, rec_str, subject, mtype, key_data))
                conn.commit()

            except Exception as e:
                errMsg = repr(e)
                send_mail('cn-tam-auto@amazon.com', [msg.get('From', "")], 'Error:' + subject, errMsg, [])
            finally:
                cursor.close()
                conn.close()

def processA9(msg,msgid,name):
    subject = ""
    for header in ["From", "To", "Subject", "Cc"]:
        value = msg.get(header, "")
        if value:
            if header == "Subject":
                value = decode_str(value)
        print("%s:%s" % (header, value))
        if header == "Subject": subject = value
    for part in msg.walk():
        # fileName = part.get_filename()
        contentType = part.get_content_type()
        if contentType.find('image/') != -1: continue
        if contentType == 'text/plain' or contentType == 'text/html':
            # 保存正文
            content = part.get_payload(decode=True)
            charset = guess_charset(part)  # or msg
            if charset:
                content = content.decode(charset)
            soup = BeautifulSoup(content, 'lxml')
            table = soup.find('table',id="dataTable")
            if table is None: continue
            table = table.find('tbody')
            if table is not None:
                line_no =0
                data = []
                for tr in table.findAll('tr'):
                    line_no += 1
                    col_no = 0
                    keywords, rank, groups = '','',''
                    for td in tr.findAll('td'):
                        col_no += 1
                        if col_no == 1:
                            keywords = td.getText().strip().replace('\r','').replace('\n','')
                        elif col_no ==3:
                            rank     = td.getText().strip()
                        elif col_no ==4:
                            groups = td.getText().strip().replace(',','')
                        elif col_no>4:
                            break
                    if keywords.find('Total:')<0:
                        data.append((keywords,rank,groups))
                if len(data) > 0:
                    conn = pymysql.connect(host=mysql_host, user=mysql_user, passwd=mysql_password,
                                           db=mysql_default_db, port=3306, local_infile=1, charset='utf8',
                                           cursorclass=pymysql.cursors.DictCursor)
                    cursor = conn.cursor()
                    rec_time = re.findall(r'\w+, (\d{1,2}) (\w+) (\d{4}) (\d{2}:\d{2}:\d{2})', msg.get("Date", ""))[0]
                    rec_str = str(datetime.datetime.strptime(
                        "%s-%s-%s %s" % (rec_time[2], months[rec_time[1].lower()], rec_time[0], rec_time[3]),
                        "%Y-%m-%d %H:%M:%S") + datetime.timedelta(hours=8))
                    cursor.execute("delete from tam_db.mail_A9 where msgid=%s" % msgid)
                    for line in data:
                        sql = "insert into tam_db.mail_A9(msgid,receive,keywords,rank,query_groups) values (%s,'%s','%s',%s,%s)" % ( (msgid, rec_str) + line )
                        cursor.execute(sql)
                    conn.commit()
                    conn.close()
            send_mail('cn-tam-auto@amazon.com', [msg.get('From', "")], 'RE:' + subject, 'A9 search data is successfully stored.\nYou are welcome!',[])

def parseAsinList(str):
    return re.findall(r'(B\w{9})',str)

def getMsgContent(msg):
    content =""
    for part in msg.walk():
        #fileName = part.get_filename()
        contentType = part.get_content_type()
        #mycode = part.get_content_charset();
        # 保存附件
        if contentType.find('image/') != -1: continue
        if contentType == 'text/plain'or contentType == 'text/html':
            # 保存正文
            content = part.get_payload(decode=True)
            charset = guess_charset(part) #or msg
            if charset:
                content = content.decode(charset)
    return content

def processSQL(msg,msgid,name):
    subject = decode_str(msg.get("Subject", ""))
    content = getMsgContent(msg)
    clearText = BeautifulSoup(content, 'lxml').get_text().strip().replace(u'\xa0', u' ')
    select_pos = clearText.lower().find('select')
    if select_pos < 0:
        send_mail('cn-tam-auto@amazon.com', [msg.get('From', "")], 'RE:' + subject,
                  'No valid SQL statement was found, it must contain the \'select\' keyword.\n\nFor example:\nselect * from tam_db.test',
                  [])
        return
    sql = clearText[select_pos:]
    end_pos = sql.find(';')
    if end_pos > 0: sql = sql[:end_pos]
    end_pos = sql.find(u'；')
    if end_pos > 0: sql = sql[:end_pos]
    sqls = [('report',sql)]
    mtype = "#SQL"
    out_path = './report_' + time.strftime("%Y-%m-%d") + '.xls'
    subject = 'RE:' + subject
    key_data = sql[:250]
    sqlQueryAndReply(msg,mtype,sqls,subject,"attach",out_path)
    key_data = key_data.replace("'", "\\'")
    return {'key_data': key_data, 'mtype': mtype}

def processHiASIN(msg,msgid,name):
    subject = decode_str(msg.get("Subject", ""))
    content = getMsgContent(msg)
    clearText = BeautifulSoup(content, 'lxml').get_text().strip().replace(u'\xa0', u' ')
    asins = parseAsinList(clearText)
    if len(asins) == 0:
        send_mail('cn-tam-auto@amazon.com', [msg.get('From', "")], 'RE:' + subject,
                  'No asin was recognized\n\nPlease make sure you have provided the valid asin list.\n\nFor example:\nB07263NW7T\nB000FA5KFA',
                  [])
        return
    in_condition = "','".join(asins)
    sql = u"select tt.asin as KASIN, ab.title_name as '书名', \
    ab.band as 'BAND', ab.onsite_date as '上线日期', ab.pubcode as 'PUBCODE', \
    tt.SNAPSHOT_DAY as '销售日',round(tt.gms/tt.pu,2) as '均价', \
    round(tt.gms,2) as '销售额', tt.pu as '销量', dgv.gv as '浏览量',concat(round(tt.pu/dgv.gv*100,1),'%%') as '转化率' , \
    kd.qb as 'KU下载量', dd.title as '促销种类' \
    from (SELECT das.asin, das.SNAPSHOT_DAY,sum(das.gms)/sum(das.pu) as asp, sum(das.gms) as gms, sum(das.pu) as pu FROM hera.daily_asin_sales das \
    where das.asin  IN ('%s') \
    group by das.asin, das.SNAPSHOT_DAY) tt \
    left join hera.ku_daily kd \
    on kd.asin = tt.asin \
    and tt.SNAPSHOT_DAY = kd.O_DAY \
    inner join hera.all_bands ab \
    on tt.asin = ab.asin \
    left join hera.deal_daily dd \
    on dd.ASIN = tt.ASIN \
    and dd.d_date  = tt.SNAPSHOT_DAY \
    left join hera.daily_gv dgv \
    on dgv.asin = tt.asin \
    and dgv.sdate = tt.SNAPSHOT_DAY \
    order by tt.asin, tt.SNAPSHOT_DAY" % in_condition
    sqls = [('report', sql)]
    mtype = "#HiASIN"
    out_path = './report_' + time.strftime("%Y-%m-%d") + '.xls'
    subject = 'RE:' + subject
    key_data = ','.join(asins)[:250]
    sqlQueryAndReply(msg,mtype, sqls, subject, "attach",out_path)
    return {'key_data': key_data, 'mtype': mtype}

def processTPNEWOS(msg,msgid,name):
    subject = decode_str(msg.get("Subject", ""))
    sql = u"select ab.*,ab.vm_name,gogo.title as deal_type, dd.deal_date as deal_date from hera.all_bands ab \
      left join (select ll.asin, max(ll.d_date) as deal_date  from hera.deal_daily ll group by ll.asin) dd  \
      on dd.ASIN = ab.asin left join hera.deal_daily gogo on gogo.asin = dd.asin and gogo.d_date = dd.deal_date \
      where ab.asin in (select yy.asin from (select ab.asin from hera.db_info di inner join hera.all_bands ab \
      on ab.isbn = di.isbn and ab.OUR_PRICE <> 0 and onsite_date >= '2017-01-01' where \
      score >=7 and di.pub_date >= '20170101' and ( rate_number>=2000 or wishes >= 2000) union all \
      SELECT ab.asin FROM hera.amazon_p_info api inner join hera.all_bands ab on ab.isbn= api.isbn  and ab.our_price <> 0 \
      and onsite_date >= '2017-01-01' where api.isbn <> '' and api.ASIN  in (select des_asin from hera.refer_table) \
      union all select ao.kasin as asin from hera.a9_on ao) as yy) and onsite_date >= '2017-01-01' order by onsite_date desc"
    sqls = [('report', sql)]
    mtype = "#TPNEWOS"
    out_path = './TAM_Pattern_NEWOS_Part.xls'
    subject = 'RE:' + subject
    key_data = ''
    sqlQueryAndReply(msg,mtype, sqls, subject, "attach",out_path)
    return {'key_data': key_data, 'mtype': mtype}

def processPriceLadder(msg,msgid,name):
    subject = decode_str(msg.get("Subject", ""))
    sql = '''
    SELECT pl.price, concat(round(ratio*100,1), '%') as pu_ratio, gms_r.gms_ratio
FROM tam_db.price_ladder pl
inner join (SELECT price, concat(round(ratio*100,1), '%') as gms_ratio
FROM tam_db.price_ladder pl
where pl.snapshot_day = (
select max(snapshot_day) from tam_db.price_ladder
)
and pl.type= 'gms'
and pl.ratio <= 0.8
order by price) gms_r
on gms_r.price = pl.price
where pl.snapshot_day = (
select max(snapshot_day) from tam_db.price_ladder
)
and pl.type= 'pu'
and pl.ratio>= 0.8
order by price
    '''
    sqls = [('report', sql)]
    mtype = "#PriceLadder"
    out_path = './PriceLadder.xls'
    subject = 'RE:' + subject
    key_data = ''
    sqlQueryAndReply(msg,mtype, sqls, subject, "attach",out_path)
    return {'key_data': key_data, 'mtype': mtype}

def sqlQueryAndReply(msg,mtype,sqls, subject,method,attach_name):
    conn = pymysql.connect(host=mysql_host, user=mysql_user, passwd=mysql_password,
                           db=mysql_default_db, port=3306, local_infile=1, charset='utf8',
                           cursorclass=pymysql.cursors.DictCursor)
    cursor = conn.cursor()
    oriSubject = decode_str(msg.get("Subject", ""))
    if subject is None or len(subject)==0:
        subject = 'RE:'+ oriSubject
    clearText = BeautifulSoup(getMsgContent(msg), 'lxml').get_text().strip().replace(u'\xa0', u' ')
    try:
        if method.lower()=='attach' or 'sharedisk' in method.lower():
            #display the data in the attachment of the mail
            workbook = xlwt.Workbook()
            # for (sname,sql) in sqls:
            param_required=False
            for sql_def in sqls:
                (sname,sql) = sql_def[:2]
                if len(sql_def)>2:
                    #(field,rep) list as inputs
                    sub_conditions=[]
                    for (field,rep) in sql_def[2]:
                        search_list = re.findall(rep,clearText)
                        if len(search_list)==0: continue
                        sub_conditions.append(field+" in ('%s')" % "','".join(search_list))
                    if len(sub_conditions)>0:
                        ins_pos = sql.upper().find('WHERE ')
                        if ins_pos>0:
                            sql = sql[:ins_pos+5] + ' ('+' or '.join(sub_conditions)+') and '+sql[ins_pos+5:]
                            sheet  = workbook.add_sheet(sname, cell_overwrite_ok=True)
                            newSheetWithSQL(sql_correct(sql), sheet, cursor)
                    else:
                        param_required=True
                else:
                    sheet = workbook.add_sheet(sname, cell_overwrite_ok=True)
                    newSheetWithSQL(sql_correct(sql), sheet, cursor)
            if param_required:
                send_mail('cn-tam-auto@amazon.com', [msg.get('From', "")], 'Error:' + oriSubject, 'Please provide the valid params.', [])
                return
            workbook.save(attach_name)
            #gzip the file if it is greater than 5M
            if os.stat(attach_name).st_size/(1024*1024) > 5:
                os.system("gzip "+attach_name)
                if os.path.exists(attach_name+'.gz'):
                    attach_name = attach_name+'.gz'

            body_text = 'Please see the attached data.'
            if mtype == "#TPNEWOS":
                body_text = '''
                       <p>See the attached data.</p>
                       <p>Please refer the <a href="https://w.amazon.com/bin/view/CN_TAM_PATTERN/">wiki</a> to get help for some features.</p>
                        '''
            if 'sharedisk' in method.lower():
                folder_name = datetime.datetime.strftime(datetime.datetime.now(), '%Y-%m-%d-%H%M')
                share_path=''
                if method[-1]==os.sep:
                    share_path = method[10:]+folder_name+os.sep
                else:
                    share_path = method[10:] + os.sep+folder_name + os.sep
                os.system("mkdir -p " + share_path)  # create it if it doesn't exist
                os.system("cp -rf " + attach_name + ' ' + share_path)  # create it if it doesn't exist
                link_path = r'\\ant\dept-as\pek02\WOL-CN' + '\\'+ share_path[9:].replace('/','\\')
                body_text = '''
                    <p>Please review the data in the following shared folder.</p>
                    <p><a href="%s">%s</a></p>
                    '''
                body_text = body_text % (link_path, link_path)
                send_mail('cn-tam-auto@amazon.com', [msg.get('From', "")], subject,body_text, [], 'html')
            else:
                send_mail('cn-tam-auto@amazon.com', [msg.get('From', "")], subject,
                      body_text, [attach_name],'html')
            if os.path.exists(attach_name):
                os.remove(attach_name)
        elif method.lower()=='table':
            #display the data in mail body as table format
            param_required = False
            mail_body=''
            list_conn=tamCommonLib.getMysqlConnect()
            list_cursor=list_conn.cursor()
            for sql_def in sqls:
                (sname, sql) = sql_def[:2]
                if len(sql_def) > 2:
                    # (field,rep) list as inputs
                    sub_conditions = []
                    for (field, rep) in sql_def[2]:
                        search_list = re.findall(rep, clearText)
                        if len(search_list) == 0: continue
                        sub_conditions.append(field + " in ('%s')" % "','".join(search_list))
                    if len(sub_conditions) > 0:
                        ins_pos = sql.upper().find('WHERE ')
                        if ins_pos > 0:
                            sql = sql[:ins_pos + 5] + ' (' + ' or '.join(sub_conditions) + ') and ' + sql[ins_pos + 5:]
                            cnt = list_cursor.execute(sql)
                            if cnt<=1000:
                                # header
                                data = [[col[0] for col in list_cursor.description]]
                                # data
                                data.extend(list_cursor.fetchall())
                                mail_body += tamCommonLib.table_html_with_rn(data, sname)
                            else:
                                mail_body +='<p>More than 1000 rows were got, so skipped.</p>'
                    else:
                        param_required = True
                else:
                    cnt = list_cursor.execute(sql)
                    if cnt <= 1000:
                        # header
                        data = [[col[0] for col in list_cursor.description]]
                        # data
                        data.extend(list_cursor.fetchall())
                        mail_body += tamCommonLib.table_html_with_rn(data, sname)
                    else:
                        mail_body += '<p>More than 1000 rows were got, so skipped.</p>'
            if param_required:
                send_mail('cn-tam-auto@amazon.com', [msg.get('From', "")], 'Error:' + oriSubject,
                          'Please provide the valid params.', [])
                return
            send_mail('cn-tam-auto@amazon.com', [msg.get('From', "")], subject, mail_body, [], 'html')

    except Exception as e:
        errMsg = repr(e)
        send_mail('cn-tam-auto@amazon.com', [msg.get('From', "")], 'Error:' + oriSubject, errMsg, [])
    finally:
        cursor.close()
        conn.close()

def checkPermission(usr,app,usr_grp):
    permission = False
    if 'all' in app['groups'] or usr in app['users']:
        permission = True
    else:
        for grp in app['groups']:
            if usr_grp.get(grp) is not None and usr in usr_grp[grp]:
                permission = True
                break
    return permission

def do_somework():
    conn = pymysql.connect(host=mysql_host, user=mysql_user, passwd=mysql_password,
                           db=mysql_default_db, port=3306, local_infile=1, charset='utf8',
                           cursorclass=pymysql.cursors.DictCursor)
    cursor = conn.cursor()
    sql='''
    update tam_db.kcode_metrics a1 inner join (
      select vendor_code as pubcode,max(tam_ops) as tam_ops 
        from tam_db.visibility where tam_ops is not null and length(tam_ops)>0 group by vendor_code) a2 on a1.digital_pubcode=a2.pubcode
     set a1.tamops = a2.tam_ops
    '''
    cursor.execute(sql)
    conn.commit()
    cursor.close()
    conn.close()


def future_process(msg,msgid):
    subject = decode_str(msg.get("Subject", ""))
    fromStr = msg.get("From", "")
    ret = re.findall(r'"(.*?)"', fromStr)
    name = ""
    usr = re.findall(r'(\w+?)@', fromStr)[0]
    if len(ret) > 0:
        name = ret[0]
        temp = name.find(',')
        if temp >= 0: name = name[temp + 1:]
    if len(name) == 0:
        name = re.findall(r'(\w+?)@', fromStr)[0]
    trust_address = False
    for add in allow_addresses:
        if msg.get("From", "").lower().find(add.lower()) >= 0:
            trust_address = True
            break
    if not trust_address or msg.get("From", "").find("cn-tam-auto@amazon.com") > 0:
        # send the warning mail to cn-kindle-mtam group
        send_mail('cn-tam-auto@amazon.com', [username + '@amazon.com'], 'Receive mail from ' + msg.get("From", ""),
                  'With below subject:\n' + subject, [])
        return

    reload(mailProcessor_config)
    configs = mailProcessor_config.apps
    is_welcome = True
    # build the user group relationship dict
    conn = pymysql.connect(host=mysql_host, user=mysql_user, passwd=mysql_password,
                           db=mysql_default_db, port=3306, local_infile=1, charset='utf8',
                           cursorclass=pymysql.cursors.DictCursor)
    cursor = conn.cursor()
    cursor.execute("select grp,usr from accesslog.usr_grp")
    usr_grp = {}
    for row in cursor.fetchall():
        if usr_grp.get(row['grp']) is None:
            usr_grp[row['grp']] = [row['usr']]
        else:
            usr_grp[row['grp']].append(row['usr'])

    for idx in range(len(configs)):
        if subject.upper().find(configs[idx]['search_str']) >= 0:
            is_welcome = False
            if checkPermission(usr, configs[idx], usr_grp):
                last_ret = {}
                if configs[idx].get('type') is None or configs[idx].get('type')[:3].upper() != 'SQL':
                    for cmd in configs[idx]['cmds']:
                        if cmd.strip().startswith('import'):
                            module_str = cmd.strip()[7:]
                            locals()[module_str] = __import__(module_str)
                        else:
                            last_ret = eval(cmd)
                else:
                    # SQL mode
                    mtype = '#' + configs[idx]['appname']
                    method = 'attach'
                    if len(configs[idx].get('type')) > 3:
                        method = configs[idx].get('type')[4:]
                    sqlQueryAndReply(msg, mtype, configs[idx]['sqls'], None, method, './report.xls')
                    last_ret = {'key_data': '', 'mtype': mtype}

                if last_ret is not None and isinstance(last_ret, dict) and last_ret.get(
                        'key_data') is not None and last_ret.get('mtype') is not None:
                    key_data = last_ret.get('key_data')
                    mtype = last_ret.get('mtype')
                    fro = re.findall(r'([a-zA-Z0-9_\-\.]+@[a-zA-Z0-9]+\.[a-zA-Z0-9]+)', msg.get('From', ""))[0]
                    rec_time = re.findall(r'\w+, (\d{1,2}) (\w+) (\d{4}) (\d{2}:\d{2}:\d{2})', msg.get("Date", ""))[0]
                    rec_str = str(datetime.datetime.strptime(
                        "%s-%s-%s %s" % (rec_time[2], months[rec_time[1].lower()], rec_time[0], rec_time[3]),
                        "%Y-%m-%d %H:%M:%S") + datetime.timedelta(hours=8))
                    cursor.execute(
                        "insert into accesslog.mail_audit(msgid,fro,receive,subject,type,key_data) values (%s,'%s','%s','%s','%s','%s')" % (
                            int(bytes.decode(msgid)), fro, rec_str, subject, mtype, key_data))
                    conn.commit()
            else:
                send_mail('cn-tam-auto@amazon.com', [msg.get('From', "")],
                          'Please request the permission first',
                          'Sorry,you don\'t have permission to access this feature.\nPlease contact TAM team to grant you the access.',
                          [])
            break
    # send out the welcome mail
    if is_welcome == True:
        cnt = 0
        trHtml = ""
        for idx in range(len(configs)):
            if configs[idx]['display'] != 'y': continue
            if checkPermission(usr, configs[idx], usr_grp):
                cnt += 1
                trHtml += "<tr>  \
                                             <td>(%s) [#%s]</td> \
                                             <td>-- %s </td> \
                                             <td><a href=\"mailto:cn-tam-auto@amazon.com?subject=[#%s]\">click me to mail</a></td> \
                                  </tr>" % (
                str(cnt), configs[idx]['appname'], configs[idx]['desc'], configs[idx]['appname'])
        send_mail('cn-tam-auto@amazon.com', [msg.get('From', "")], 'Welcome to use TAM AUTO mail system',
                  helpContent % (name, trHtml), [], 'html')

    cursor.close()
    conn.close()

latest_run = {}
def do_schedulers(threadPool):
    st = time.localtime(time.time())
    dow = st.tm_wday + 1
    cur_min = st.tm_hour * 60 + st.tm_min
    reload(mailProcessor_config)
    for scheduler in mailProcessor_config.schedulers:
        if latest_run.get(scheduler['name']) is not None and time.time() - latest_run.get(scheduler['name'])<=3600:
            #if has run it in one hour, then ignore it
            continue
        times = scheduler['time'].split(':')
        sch_min = int(times[0])*60
        if len(times)>1: sch_min +=int(times[1])
        days=scheduler['day_of_week'].split(',')
        get_in_window = False
        if cur_min>=sch_min and cur_min<=sch_min+10:
            for dd in days:
                dds=dd.split('-')
                if (len(dds)>1 and dow>=int(dds[0]) and dow<=int(dds[1])) or (len(dds)==1 and dow==int(dds[0])):
                    get_in_window = True
                    break
        if get_in_window:
            user_list = []
            conn = tamCommonLib.getMysqlConnect()
            cursor = conn.cursor()
            if len(scheduler['groups'])>0:
                cursor.execute("select distinct usr from accesslog.usr_grp where grp in ('%s')" % "','".join(scheduler['groups']))
                user_list.extend([dd[0] if '@' in dd[0] else dd[0] + '@amazon.com' for dd in cursor.fetchall()])
            if len(scheduler['users'])>0:
                for user in scheduler['users']:
                    if '@' not in user: user +='@amazon.com'
                    if user_list.count(user)==0: user_list.append(user)
            mail_body =''
            cnt =0
            wb =''
            for sql_def in scheduler['sqls']:
                (sname, sql) = sql_def[:2]
                append_rn = True
                if len(sql_def) > 2 and sql_def[2]==False:
                    append_rn =False
                cursor.execute(sql_correct(sql))
                if scheduler['publisher_type']=='table':
                    #header
                    data = [[col[0] for col in cursor.description]]
                    #data
                    data.extend(cursor.fetchall())
                    if append_rn==True:
                        mail_body +=tamCommonLib.table_html_with_rn(data,sname)+'<p></p>'
                    else:
                        mail_body += tamCommonLib.table_html(data, sname) + '<p></p>'
                elif scheduler['publisher_type']=='attach':
                    if cnt ==0:
                        wb = Workbook()
                    mail_body ='Please see the attached data'
                    ns = wb.worksheets[cnt]
                    ns.title = 'report%s' % str(cnt+1)
                    ns.append([col[0] for col in cursor.description])
                    for row in cursor.fetchall():
                        ns.append(row)
                cnt +=1
            if scheduler['publisher_type']=='attach':
                ff = './report.xlsx'
                wb.save(ff)
                send_mail('cn-tam-auto@amazon.com', user_list,scheduler['subject'],mail_body, [ff], 'html')
                os.remove(ff)
            else:
                send_mail('cn-tam-auto@amazon.com', user_list, scheduler['subject'], mail_body, [], 'html')
            latest_run[scheduler['name']] = time.time()
            cursor.close()
            conn.close()


username = input("username:")
password = getpass.getpass('password:')
mailserver = imaplib.IMAP4_SSL('ballard.amazon.com', 1993)
mailserver.login(username+"@amazon.com/cn-tam-auto", password)

looper = 0
pool = ThreadPoolExecutor(max_workers=5)
while True:
    try:
        if looper % (2*60) == 0:  #get into this every 60 minutes
            do_somework()
        if looper % (2*5)  == 0:  #get into this every 5 minutes to check the schedulers
            do_schedulers(pool)
        looper += 1
        status, count = mailserver.select('Inbox')
        ttype, data = mailserver.search(None, 'unseen')
        msgList = data[0].split()
        for msgid in msgList:
            ttype, ddata = mailserver.fetch(msgid,'(RFC822)')
            msg=email.message_from_string(ddata[0][1].decode('utf-8'))
            pool.submit(future_process,msg,msgid)
            mailserver.store(msgid,'+FLAGS','\\seen')
        print("sleep for 30 seconds at %s......" % str(datetime.datetime.now()))
        time.sleep(30)
        print("wake up to run again at %s......" % str(datetime.datetime.now()))
    except Exception as e:
        print(repr(e))
        if repr(e).find('socket error')>= 0:
            send_mail('cn-tam-auto@amazon.com', ['yongqis@amazon.com'], 'Socket Error happended for mailProcessor.py program.', 'Please resolve it ASAP.\n', [])
            time.sleep(1200)
            #relogin to the mail server
            mailserver = imaplib.IMAP4_SSL('ballard.amazon.com', 1993)
            mailserver.login(username + "@amazon.com/cn-tam-auto", password)
    finally:
        running = 2
