#!/usr/bin/python3
#coding=utf-8

import imaplib
import email
from email.parser import Parser
from email.header import decode_header
from email.utils import parseaddr
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.utils import COMMASPACE,formatdate
from email import encoders

from bs4 import BeautifulSoup
import pymysql
import importlib
import sys
import logging
import getpass
import time
import xlrd
import os
import xlwt
import nltk
import unicodedata
import re
import datetime

import xlrd
import datetime
from openpyxl import Workbook

## comment variables ##########################################
## time zone
mysql_host = 'ud094661c879c59fa6e9e'
mysql_user = '***'
mysql_password = '***'
mysql_default_db = '***'
share_path = '/mnt/dept1/Content-Demand/CLC/Billing/Raw_Data/'
display_path = '\\\\ant\\dept-as\\pek02\\dept1\\Content-Demand\\CLC\\Billing\\Raw_Data\\'
## mysql configuration
importlib.reload(sys)

###############################################################

# 邮件的Subject或者Email中包含的名字都是经过编码后的str，要正常显示，就必须decode
def decode_str(s):
    value, charset = decode_header(s)[0]
    # decode_header()返回一个list，因为像Cc、Bcc这样的字段可能包含多个邮件地址，所以解析出来的会有多个元素。上面的代码我们偷了个懒，只取了第一个元素。
    if charset:
        value = value.decode(charset)
    return value

# 文本邮件的内容也是str，还需要检测编码，否则，非UTF-8编码的邮件都无法正常显示
def guess_charset(msg):
    charset = msg.get_charset()
    if charset is None:
        content_type = msg.get('Content-Type', '').lower()
        pos = content_type.find('charset=')
        if pos >= 0:
            charset = content_type[pos + 8:].strip()
    return charset

def send_mail(fro, to, subject, text, files=[], mtype='plain'):
    #assert type(server) == dict
    #assert type(to) == list
    #assert type(files) == list

    msg = MIMEMultipart()
    msg['From'] = fro
    msg['Subject'] = subject
    msg['To'] = ','.join(to)  # COMMASPACE==', '
    #msg['Date'] = formatdate(localtime=True)
    msg.attach(MIMEText(text,mtype,'utf-8'))

    for file in files:
        part = MIMEBase('application', 'octet-stream')  # 'octet-stream': binary data
        part.set_payload(open(file, 'rb').read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', 'attachment; filename="%s"' % os.path.basename(file))
        msg.attach(part)

    import smtplib
    smtp = smtplib.SMTP('mail-relay.amazon.com')
    #smtp.login(server['user'], server['passwd'])
    smtp.sendmail(fro, to, msg.as_string())
    smtp.close()

def newSheetWithSQL(sql,sheet,cursor,accumulate_data):
    results = []
    looper = 0
    while True:
        looper +=1
        try:
            cursor.execute(sql)
            results = cursor.fetchall()
            break
        except Exception as e:
            print(repr(e))
            time.sleep(10)
            if looper==100: break
            if looper%10==0:
                cursor=pymysql.connect(host=mysql_host, user=mysql_user, passwd=mysql_password,
                                db=mysql_default_db, port=3306, local_infile=1, charset='utf8',
                                cursorclass=pymysql.cursors.DictCursor).cursor()

    fields = cursor.description
    # 写上字段信息
    #for field in range(len(fields)):
    #    _=sheet.cell(row=1, column=field+1, value=str(fields[field][0]))
    if accumulate_data is not None and len(accumulate_data)==0: accumulate_data.append([fields[idx][0] for idx in range(len(fields))])
    # 获取并写入数据段信息
    row = 1
    col = 0
    for irow in range(1, len(results) + 1):
        rowdata=[]
        for col in range(0, len(fields)):
            if results[irow - 1].get(fields[col][0]) is not None:
                if isinstance(results[irow - 1].get(fields[col][0]), datetime.date):
                    #sheet.write(row, col, results[row - 1].get(fields[col][0]), dateFormat)
                    #_=sheet.cell(row=irow+1, column=col+1, value=str(results[irow - 1].get(fields[col][0])))
                    rowdata.append(str(results[irow - 1].get(fields[col][0])))
                else:
                    #_=sheet.cell(row=irow+1, column=col+1, value=results[irow - 1].get(fields[col][0]))
                    rowdata.append(results[irow - 1].get(fields[col][0]))
            else:
                rowdata.append(None)
        if accumulate_data is not None and results[irow - 1].get('vm')!='Total':
            accumulate_data.append(rowdata)

def processBill(msg,msgid,name):
    models = ['vendor-fund', 'co-fund-fixed', 'co-fund-ordered']
    subject = ""
    promotions = []
    if msg is not None:
        for header in ["From", "To", "Subject", "Cc"]:
            value = msg.get(header, "")
            if value:
                if header == "Subject":
                    value = decode_str(value)
            print("%s:%s" % (header, value))
            if header == "Subject": subject = value
        for part in msg.walk():
            fileName = part.get_filename()
            contentType = part.get_content_type()
            if contentType.find('image/') != -1: continue
            if contentType == 'text/plain' or contentType == 'text/html':
                # 保存正文
                content = part.get_payload(decode=True)
                charset = guess_charset(part)  # or msg
                if charset:
                    content = content.decode(charset)
                soup = BeautifulSoup(content, 'lxml')
                # save the metadata
                table = soup.find('table')
                if table is not None:
                    for tr in table.findAll('tr'):
                        col_no = 0
                        prom_id, model, cost_rate = "", "", ""
                        for td in tr.findAll('td'):
                            col_no += 1
                            if col_no==1:
                                prom_id = td.getText().strip()
                            elif col_no==2:
                                model = td.getText().strip()
                            elif col_no==3:
                                cost_rate = td.getText().strip()
                        if len(prom_id)>0 and len(model)>0 and model.lower() in models:
                            promotions.append([prom_id,model.lower(),cost_rate])
            if fileName:
                data = part.get_payload(decode=True)
                fname = decode_str(fileName)
                fEx = open("%s" % (fname), 'wb')
                fEx.write(data)
                fEx.close()
                if fname.find('.xls') >= 0:
                    data = xlrd.open_workbook(fname)
                    table = data.sheet_by_index(0)
                    nrows = table.nrows
                    ncols = table.ncols
                    prom_id, model, cost_rate = "", "", ""
                    for row_id in range(nrows):
                        if ncols >= 1: prom_id = str(table.cell(row_id,0).value).strip()
                        if ncols >= 2: model = str(table.cell(row_id, 1).value).strip()
                        if ncols >= 3: cost_rate = str(table.cell(row_id, 2).value).strip()
                        if len(prom_id)>0 and len(model)>0 and model.lower() in models:
                            if len(re.findall(r'(\d+?)\.',prom_id))>0: prom_id = re.findall(r'(\d+?)\.',prom_id)[0]
                            promotions.append([prom_id,model.lower(),cost_rate])
                if os.path.exists(fname): os.remove(fname)

    if len(promotions)==0:
        send_mail('cn-tam-auto@amazon.com', [msg.get('From', "")], 'No valid promotion data was received',"Please provide valid data.", [])
        return

    conn = pymysql.connect(host=mysql_host, user=mysql_user, passwd=mysql_password,
                           db=mysql_default_db, port=3306, local_infile=1, charset='utf8',
                           cursorclass=pymysql.cursors.DictCursor)
    cursor = conn.cursor()

    folder_name = datetime.datetime.strftime(datetime.datetime.now(),'%Y-%m-%d-%H%M')
    all_summary =[]
    all_details =[]
    for prom in promotions:
        sql,sql_d="",""
        if prom[1]==models[0]:
            temp ='''
             select promo_id,vendor_code as pub,max(km.vm_name_s) as vm,count(1) as order_cnt,sum(1-is_return) as valid_cnt,sum((1-is_return)*our_price_with_tax_mp_currency) as  Revenue_including_coupon,
              sum((1-is_return)*(0 - discount_amount - discount_amount_tax)) as coupon_amount,sum((1-is_return)*item_cost) as original_cost_amount, sum((1-is_return)*(item_cost+discount_amount+discount_amount_tax)) as cost_amount_payable,
              sum((1-is_return)*(0 - discount_amount - discount_amount_tax)) as cost_adjustment_amount from tam_db.promotion_billing pb
              left join kcode_metrics km on km.digital_pubcode=pb.vendor_code
              where is_charged='Y' replacement_here group by promo_id,vendor_code
               '''
            sql = temp + " union  select '','','Total',sum(order_cnt),sum(valid_cnt),sum(Revenue_including_coupon),sum(coupon_amount),sum(original_cost_amount),sum(cost_amount_payable),sum(cost_adjustment_amount) from (" +temp+") aa"
            sql_d ="select promo_id,promo_description,asin,vendor_code as pub,order_id,order_day_local as order_date,our_price_with_tax_mp_currency as Revenue_including_coupon,0 - discount_amount - discount_amount_tax as coupon_amount,\
               item_cost as original_cost_amount, item_cost+discount_amount+discount_amount_tax as cost_amount_payable,0 - discount_amount - discount_amount_tax as cost_adjustment_amount,is_return \
               from tam_db.v_promotion_billing where is_charged='Y' replacement_here order by promo_id,vendor_code,order_day_local"
        elif prom[1]==models[1]:
            tt = re.findall(r'^([0-9]*\.?[0-9]+|[0-9]+\.?[0-9]*)%$', prom[2])
            rate = 0
            if len(tt) == 0:
                rate = float(prom[2])
            else:
                rate = float(tt[0]) * 0.01
            temp = '''select promo_id,vendor_code as pub,max(km.vm_name_s) as vm,count(1) as order_cnt,sum(1-is_return) as valid_cnt,sum((1-is_return)*our_price_with_tax_mp_currency) as Revenue_including_coupon,
                      sum((1-is_return)* (0 - discount_amount - discount_amount_tax)) as coupon_amount,sum((1-is_return)*item_cost) as original_cost_amount, round(%s * sum((1-is_return)*(our_price_with_tax_mp_currency+discount_amount+discount_amount_tax)),2) as cost_amount_payable,
                       round(sum((1-is_return)*item_cost) - %s * sum((1-is_return)*(our_price_with_tax_mp_currency+discount_amount+discount_amount_tax)),2) as cost_adjustment_amount
                       from tam_db.promotion_billing pb left join kcode_metrics km on km.digital_pubcode=pb.vendor_code
                      where is_charged='Y' replacement_here group by promo_id,vendor_code
                   '''
            sql = temp + " union  select '','','Total',sum(order_cnt),sum(valid_cnt),sum(Revenue_including_coupon),sum(coupon_amount),sum(original_cost_amount),sum(cost_amount_payable),sum(cost_adjustment_amount) from (" + temp + ") aa"
            sql = sql % (str(rate),str(rate),str(rate),str(rate))
            sql_d = "select promo_id,promo_description,asin,vendor_code as pub,order_id,order_day_local as order_date,our_price_with_tax_mp_currency as Revenue_including_coupon,0 - discount_amount - discount_amount_tax as coupon_amount,\
                           item_cost as original_cost_amount, round(%s * (our_price_with_tax_mp_currency+discount_amount+discount_amount_tax),2) as cost_amount_payable,round(item_cost - %s * (our_price_with_tax_mp_currency+discount_amount+discount_amount_tax),2) as cost_adjustment_amount,is_return \
                           from tam_db.v_promotion_billing where is_charged='Y' replacement_here order by promo_id,vendor_code,order_day_local" % (str(rate),str(rate))
        elif prom[1]==models[2]:
            temp = '''
             select promo_id,vendor_code as pub,max(km.vm_name_s) as vm,count(1) as order_cnt,sum(1-is_return) as valid_cnt,sum((1-is_return)*our_price_with_tax_mp_currency) as Revenue_including_coupon,
                       sum((1-is_return)*(0 - discount_amount - discount_amount_tax)) as coupon_amount,sum((1-is_return)*item_cost) as original_cost_amount, round(sum((1-is_return)*item_cost/our_price_with_tax_mp_currency*(our_price_with_tax_mp_currency+discount_amount+discount_amount_tax)),2) as cost_amount_payable,
                        round(sum((1-is_return)*item_cost) - sum( (1-is_return)*item_cost/our_price_with_tax_mp_currency * (our_price_with_tax_mp_currency+discount_amount+discount_amount_tax)),2) as cost_adjustment_amount 
                       from tam_db.promotion_billing pb left join kcode_metrics km on km.digital_pubcode=pb.vendor_code
                  where is_charged='Y' replacement_here group by promo_id,vendor_code
                  '''
            sql = temp + " union  select '','','Total',sum(order_cnt),sum(valid_cnt),sum(Revenue_including_coupon),sum(coupon_amount),sum(original_cost_amount),sum(cost_amount_payable),sum(cost_adjustment_amount) from (" + temp + ") aa"
            sql_d = "select promo_id,promo_description,asin,vendor_code as pub,order_id,order_day_local as order_date,our_price_with_tax_mp_currency as Revenue_including_coupon,0 - discount_amount - discount_amount_tax as coupon_amount,\
                        item_cost as original_cost_amount, round(item_cost/our_price_with_tax_mp_currency * (our_price_with_tax_mp_currency+discount_amount+discount_amount_tax),2) as cost_amount_payable,round(item_cost - item_cost/our_price_with_tax_mp_currency * (our_price_with_tax_mp_currency+discount_amount+discount_amount_tax),2) as cost_adjustment_amount,is_return \
                        from tam_db.v_promotion_billing where is_charged='Y' replacement_here order by promo_id,vendor_code,order_day_local"
        else:
            sql= 'nothing'
            sql_d='nothing'

        #file_path = share_path + folder_name + os.sep + prom[0] + os.sep
        #os.system("mkdir -p " +file_path) #create it if it doesn't exist
        #workbook = Workbook()
        #sheet = workbook.add_sheet("Summary", cell_overwrite_ok=True)
        #sheet = workbook.worksheets[0]
        #sheet.title='Summary'
        newSheetWithSQL(sql.replace('replacement_here','and promo_id=%s'%prom[0]), None, cursor,all_summary)
        #sheet = workbook.add_sheet("Detail", cell_overwrite_ok=True)
        #sheet = workbook.create_sheet(title="Detail")
        newSheetWithSQL(sql_d.replace('replacement_here','and promo_id=%s'%prom[0]), None, cursor,all_details)
        #ff_name = file_path + "%s_billing_result.xlsx" % prom[0]
        #workbook.save(ff_name)

        cursor.execute("select vendor_code,max(km.vm_name_s) as vm from tam_db.promotion_billing pb \
                        left join kcode_metrics km on km.digital_pubcode=pb.vendor_code \
                         where promo_id=%s and is_charged='Y' group by vendor_code" % prom[0])
        for row  in cursor.fetchall():
            #workbook = Workbook()
            condition = "and promo_id=%s and vendor_code='%s'" % (prom[0],row['vendor_code'])
            #sheet = workbook.add_sheet("Summary", cell_overwrite_ok=True)
            #sheet = workbook.worksheets[0]
            #sheet.title = 'Summary'
            newSheetWithSQL(sql.replace('replacement_here', condition), None, cursor,None)
            #sheet = workbook.add_sheet("Detail", cell_overwrite_ok=True)
            #sheet = workbook.create_sheet(title="Detail")
            newSheetWithSQL(sql_d.replace('replacement_here', condition), None, cursor,None)
            #ff_name = file_path + '%s_%s_%s_Billing.xlsx'% (prom[0],row['vendor_code'],row['vm'])
            #workbook.save(ff_name)

    #all promotions data
    if len(promotions)>0 and len(all_summary)>0:
        wb = Workbook()
        ns = wb.worksheets[0]
        ns.title='Summary'
        sum3,sum4,sum5,sum6,sum7,sum8,sum9 = 0,0,0,0,0,0,0
        for idx in range(len(all_summary)):
            ns.append(all_summary[idx])
            if idx!=0:
                if all_summary[idx][3] is not None: sum3 += all_summary[idx][3]
                if all_summary[idx][4] is not None: sum4 += all_summary[idx][4]
                if all_summary[idx][5] is not None: sum5 += all_summary[idx][5]
                if all_summary[idx][6] is not None: sum6 += all_summary[idx][6]
                if all_summary[idx][7] is not None: sum7 += all_summary[idx][7]
                if all_summary[idx][8] is not None: sum8 += all_summary[idx][8]
                if all_summary[idx][9] is not None: sum9 += all_summary[idx][9]
        ns.append(['','','Total',sum3,sum4,sum5,sum6,sum7,sum8,sum9])
        #sheet.write(row, col, str(results[row - 1].get(fields[col][0])))
        ns = wb.create_sheet('Detail')
        for row in all_details:
            ns.append(row)
        os.system("mkdir -p " + share_path + folder_name)
        ff_name = share_path + folder_name + os.sep + 'all_billing_result.xlsx'
        wb.save(ff_name)

    link_path = display_path + folder_name
    mailBody = '''
    <p>Please review the data in the following shared folder.</p>
    <p><a href="%s">%s</a></p>
    '''
    mailBody = mailBody % (link_path,link_path)
    if msg is not None:
        send_mail('cn-tam-auto@amazon.com', [msg.get('From', "")], 'Promotion Billing data', mailBody,[],'html')
    else:
        send_mail('cn-tam-auto@amazon.com', ['yongqis@amazon.com'], 'Promotion Billing data', mailBody,[],'html')

    return {'key_data': ','.join([promotions[idx][0] for idx in range(len(promotions))]), 'mtype': '#PromoBill'}


if __name__ == '__main__':
    processBill(None,'','')
